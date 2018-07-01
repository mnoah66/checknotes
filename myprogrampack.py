from tkinter import *
import openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta, date, time
from tkinter.font import Font
from tkinter import filedialog  


import csv
 
window = Tk()
 
window.title("Welcome to LikeGeeks app")
 
window.geometry('600x600')
 
# KeyWords
lblKeywords = Label(window, text="FLAG THE FOLLOWING WORDS")
lblKeywords.pack(fill=X,padx=50)
txt = Entry(window,width=30)
txt.insert(END, '')
txt.pack(padx=5,pady=2)

# Odd duration Greater Than
var =IntVar()
var.set(360)
lblGreaterThan = Label(window, text="FLAG DURATIONS GREATER THAN OR EQUAL TO:")
lblGreaterThan.pack(fill=X,padx=50)
spinGreaterThan = Spinbox(window, from_=0, to=1000, width=5,  font=Font(family='Helvetica', size=12, weight='bold'),textvariable=var)
spinGreaterThan.pack(padx=5,pady=2)

# Odd duration less than
var =IntVar()
var.set(30)
lblLessThan = Label(window, text="FLAG DURATIONS LESS THAN OR EQUAL TO:")
lblLessThan.pack(fill=X,padx=50)
spinLessThan = Spinbox(window, from_=0, to=1000, width=5,  font=Font(family='Helvetica', size=12, weight='bold'),textvariable=var)
spinLessThan.pack(padx=5,pady=2)

# Note length
var =IntVar()
var.set(75)
lblShortNote = Label(window, text="FLAG NOTE LENGTH LESS THAN")
lblShortNote.pack(fill=X,padx=50)
spinShortNote = Spinbox(window, from_=0, to=1000, width=5,  font=Font(family='Helvetica', size=12, weight='bold'),textvariable=var)
spinShortNote.pack(padx=5,pady=2)



# Weird Times - Start Time After
lblHour = Label(window, text="FLAG NOTES WITH START TIME AFTER:")
lblHour.pack(padx=50,pady=2)

spinHourAfter = Spinbox(window, values=("06","07","08","09","10","11","12"), font=Font(family='Helvetica', size=12, weight='bold'), width=5, state='readonly')
spinHourAfter.pack(padx=5)
spinMinAfter = Spinbox(window, values=("00","01",
"02",
"03",
"04",
"05",
"06",
"07",
"08",
"09",
"10",
"11",
"12",
"13",
"14",
"15",
"16",
"17",
"18",
"19",
"20",
"21",
"22",
"23",
"24",
"25",
"26",
"27",
"28",
"29",
"30",
"31",
"32",
"33",
"34",
"35",
"36",
"37",
"38",
"39",
"40",
"41",
"42",
"43",
"44",
"45",
"46",
"47",
"48",
"49",
"50",
"51",
"52",
"53",
"54",
"55",
"56",
"57",
"58",
"59",
), font=Font(family='Helvetica', size=12, weight='bold'), width=5, state='readonly')
spinMinAfter.pack(padx=5)
spinAMPMafter = Spinbox(window, values=("AM","PM"), font=Font(family='Helvetica', size=12, weight='bold'), width=5, state='readonly')
spinAMPMafter.pack(padx=5)

# Weird Times - Start Time BEFORE
lblHourBefore = Label(window, text="FLAG NOTES WITH TIME BEFORE:")
lblHourBefore.pack(padx=50)

spinHourBefore = Spinbox(window, values=("09","10","11","12"), font=Font(family='Helvetica', size=12, weight='bold'), width=5, state='readonly')
spinHourBefore.pack(padx=5)

spinMinBefore = Spinbox(window, values=("00","01",
"02",
"03",
"04",
"05",
"06",
"07",
"08",
"09",
"10",
"11",
"12",
"13",
"14",
"15",
"16",
"17",
"18",
"19",
"20",
"21",
"22",
"23",
"24",
"25",
"26",
"27",
"28",
"29",
"30",
"31",
"32",
"33",
"34",
"35",
"36",
"37",
"38",
"39",
"40",
"41",
"42",
"43",
"44",
"45",
"46",
"47",
"48",
"49",
"50",
"51",
"52",
"53",
"54",
"55",
"56",
"57",
"58",
"59",
), font=Font(family='Helvetica', size=12, weight='bold'), width=5, state='readonly')
spinMinBefore.pack(padx=5)

spinAMPMbefore = Spinbox(window, values=("AM","PM"), font=Font(family='Helvetica', size=12, weight='bold'), width=5, state='readonly')
spinAMPMbefore.pack(padx=5)





def file_choose():
    global file_name
    file_name = filedialog.askopenfilename() 
    lblfile = Label(window, text=file_name)
    lblfile.configure(text=file_name)
    lblfile.pack(fill=X,padx=50)  


buttonFile = Button(window, text="Choose File", command=file_choose)
buttonFile.pack(fill=X,padx=50,pady=50)






# Args to Pass to main function when button clicked
keywords = txt.get()
my_list = keywords.split(",")
greaterthan = int(spinGreaterThan.get())
lessthan = int(spinLessThan.get())
notelength = int(spinShortNote.get())
startTimeAfter = spinHourAfter.get() + ":" + spinMinAfter.get() + " " + spinAMPMafter.get()
startTimeBefore = spinHourBefore.get() + ":" + spinMinBefore.get() + " " + spinAMPMbefore.get()

def convert24(str1):
    # Checking if last two elements of time
    # is AM and first two elements are 12
    if str1[-2:] == "AM" and str1[:2] == "12":
        hour = "0"
        minute = str1[3:-3]
        return hour, minute   
    # remove the AM    
    elif str1[-2:] == "AM":
        if str1[0] == "0":
            hour = str1[1]
        else:
            hour = str1[:2]
        if str1[3] == "0":
            minute = str1[4]
        else:
            minute = str1[3:-3]
        
        #minute =  str1[3:-3]
        return hour, minute
     
    # Checking if last two elements of time
    # is PM and first two elements are 12   
    elif str1[-2:] == "PM" and str1[:2] == "12":
        if str1[0] == "0":
            hour = str1[1]
        else:
            hour = str1[:2]
        if str1[3] == "0":
            minute = str1[4]
        else:
            minute = str1[3:-3]
        return hour, minute
         
    else:
         
        # add 12 to hours and remove PM
        hour = int(str1[:2]) + 12
        
        if str1[3] == "0":
            minute = str1[4]
        else:
            minute = str1[3:-3]

        return hour, minute



def flaggedWords(ws, outWrite, my_list):

    '''Finds keywords in row of data, throws in list'''
    for row in ws:
        d = row[0] # The note
        e = row[1] # The name of the individual
        f = row[2] # Contact date
        g = row[3] # Program
        h = row[4] # Start time
        i = row[5] # end time
        j = row[6] # duration
        k = row[7] # Note writer
        foundWords = []
        if d.value:
            for w in sorted(my_list):
                if w.lower() in str(d.value).lower():
                    foundWords.append(w)
            if len(foundWords) > 0:
                note = ''
                for l in foundWords:
                    left,sep,right = d.value.lower().partition(l)
                    note = note + left[-70:] + sep.upper() + right[:70] + ';'
                forCSV = ','.join(foundWords).upper()
                outWrite.writerow([forCSV, e.value,
                	str(h.value.strftime('%I:%M%p')),
                	str(i.value.strftime('%I:%M%p')), 
                	str(f.value.strftime('%m/%d/%Y')), 
                	note, 
                	g.value, 
                	j.value, 
                	k.value])
    
def oddDuration(ws, outWrite, greaterthan, lessthan):
    for row in ws:
        d = row[0] # The note
        e = row[1] # The name of the individual
        f = row[2] # Contact date
        g = row[3] # Program
        h = row[4] # Start time
        i = row[5] # end time
        j = row[6] # duration
        k = row[7] # Note writer
        if j.value:
            if j.value < lessthan or j.value > greaterthan:
                note = d.value.split('.')
                d = '.'.join(note[1:3]).lstrip() + ' [. . .] ' + d.value[-100:]
                outWrite.writerow(['DURATION', e.value,
                	str(h.value.strftime('%I:%M%p')),
                	str(i.value.strftime('%I:%M%p')), 
                	str(f.value.strftime('%m/%d/%Y')), 
                	d, 
                	g.value, 
                	j.value, 
                	k.value])

def shortNote(ws, outWrite, notelenght):
    for row in ws:
        d = row[0] # The note
        e = row[1] # The name of the individual
        f = row[2] # Contact date
        g = row[3] # Program
        h = row[4] # Start time
        i = row[5] # end time
        j = row[6] # duration
        k = row[7] # Note writer
        if d.value:
            if len(d.value) < 75:
                outWrite.writerow(['SHORT NOTE', e.value,
                	str(h.value.strftime('%I:%M%p')),
                	str(i.value.strftime('%I:%M%p')), 
                	str(f.value.strftime('%m/%d/%Y')), 
                	d.value, 
                	g.value, 
                	j.value, 
                	k.value])

def oddTimes(ws, outWrite, startTimeAfter, startTimeBefore):
    print(startTimeAfter)
    for row in ws:
        tup1 = convert24(startTimeAfter)
        print(tup1)
        startAfterHour = tup1[0]
        startAfterMinute=tup1[1]
        print(startAfterHour)
        #startAfter = convert24(startTimeAfter)
        d = row[0] # The note
        e = row[1] # The name of the individual
        f = row[2] # Contact date
        g = row[3] # Program
        h = row[4] # Start time
        i = row[5] # end time
        j = row[6] # duration
        k = row[7] # Note writer
        if h.value:
            note = d.value.split('.')
            d = '.'.join(note[1:3]).lstrip() + ' [. . .] ' + d.value[-100:]
            try:
                if h.value > time(startAfterHour, startAfterMinute):
                    outWrite.writerow(["START TIME", e.value,
                	str(h.value.strftime('%I:%M%p')),
                	str(i.value.strftime('%I:%M%p')), 
                	str(f.value.strftime('%m/%d/%Y')), 
                	note, 
                	g.value, 
                	j.value, 
                	k.value])
            except (TypeError):
            	outWrite.writerow(["12AM or Other", e.value,
                	str(h.value.strftime('%I:%M%p')),
                	str(i.value.strftime('%I:%M%p')), 
                	str(f.value.strftime('%m/%d/%Y')), 
                	note, 
                	g.value, 
                	j.value, 
                	k.value])
               

'''
def csvWritee(title, e,h,i,f,d,g,j,k):
    title = title
    e = e.value
    h = str(h.value.strftime('%I:%M%p'))
    i = str(i.value.strftime('%I:%M%p'))
    f = str(f.value.strftime('%m/%d/%Y'))
    d = d
    g = str(g.value)
    j = str(j.value)
    k = str(k.value)
    outputWriter.writerow([title, e, h, i, f, d, g, j, k])
'''

def beginScan(my_list, greaterthan, lessthan, notelength, file_name, startTimeAfter, startTimeBefore):
	from tkinter import filedialog
	import tkinter as tk
	root = tk.Tk()
	root.withdraw()
	


	#trngfile =  openpyxl.load_workbook(r'Book1.xlsx', read_only=True)
	trngfile =  openpyxl.load_workbook(file_name, read_only=True)
	ws = trngfile['Sheet1']
	outputFile = open(r'DAYHAB_Audit on - ' + str(datetime.now().date()) + '.csv', 'w', newline='')
	outputWriter = csv.writer(outputFile)
	outputWriter.writerow(['Flagged Word/Phrase', 'Individual', 'StartTime', 'EndTime', 'Date', 'Excerpt', 'Program', 'Duration','Staff', 'Audit Comments'])
	
	flaggedWords(ws, outputWriter, my_list)
	oddDuration(ws, outputWriter, greaterthan, lessthan)
	shortNote(ws, outputWriter, notelength)
	oddTimes(ws, outputWriter, startTimeAfter, startTimeBefore)
	outputFile.close()


buttonRunMain = Button(window, text="Run Program", command= lambda: beginScan(my_list, greaterthan, lessthan, notelength, file_name, startTimeAfter, startTimeBefore))
buttonRunMain.pack(fill=X,padx=50,pady=50)

window.mainloop()