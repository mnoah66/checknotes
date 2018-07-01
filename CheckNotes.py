import tkinter as tk
from tkinter.font import Font
from tkinter import filedialog,messagebox
import openpyxl
from openpyxl import load_workbook
from datetime import timedelta, date, time
import datetime
import csv
import io
import os 
from tkinter import ttk

class SampleApp(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)

        self.title("Check My Notes")
 
        #self.geometry('600x600')

        # -------------------- MAIN GUI -----------------------------

        self.labelKeywords = tk.Label(self, text="FLAG THE FOLLOWING WORDS (Separate with comma)",font=Font(family='Arial', size=11))
        self.labelKeywords.pack(fill=tk.X,padx=50,pady=5)
        self.entryKeywords = tk.Entry(self,font=Font(family='Arial', size=11))
        self.entryKeywords.pack(fill=tk.X,padx=50, pady=2)

     
        self.labelDurationsGreater = tk.Label(self, text="FLAG TOTAL MINUTES GREATER OR EQUAL TO",font=Font(family='Arial', size=11))
        self.labelDurationsGreater.pack(fill=tk.X, padx=50)
        self.spinDurationsGreater = tk.Entry(self, width=5,  font=Font(family='Helvetica', size=12))
        self.spinDurationsGreater.pack(padx=5,pady=2)
        
        
        
        self.labelDurationsLess = tk.Label(self, text="FLAG TOTAL MINUTES LESS OR EQUAL TO",font=Font(family='Arial', size=11))
        self.labelDurationsLess.pack(fill=tk.X, padx=50)
        self.spinDurationsLess = tk.Entry(self, width=5,  font=Font(family='Helvetica', size=12))
        self.spinDurationsLess.pack(padx=5,pady=2)

        
        self.labelNoteLength = tk.Label(self, text="FLAG NOTE LENGTH (CHARACTERS) LESS THAN",font=Font(family='Arial', size=11))
        self.labelNoteLength.pack(fill=tk.X, padx=50)
        self.spinNoteLength = tk.Entry(self, width=5,  font=Font(family='Helvetica', size=12))
        self.spinNoteLength.pack(padx=5,pady=2)


        self.labelStartAfter = tk.Label(self, text="FLAG NOTE WITH START TIME AFTER",font=Font(family='Arial', size=11))
        self.labelStartAfter.pack(fill=tk.X,padx=50)
        self.spinHourAfter = tk.Spinbox(self, values=("","01","02","03","04","05","06","07","08","09","10","11","12"),font=Font(family='Helvetica', size=12), width=5,readonlybackground='white')
        self.spinHourAfter.pack(padx=5)
        self.spinMinAfter = tk.Spinbox(self, values=("","01",
        "02",
        "03",
        "04",
        "05",
        "06",
        "07",
        "08",
        "09",
        "10",
        "11",
        "12",
        "13",
        "14",
        "15",
        "16",
        "17",
        "18",
        "19",
        "20",
        "21",
        "22",
        "23",
        "24",
        "25",
        "26",
        "27",
        "28",
        "29",
        "30",
        "31",
        "32",
        "33",
        "34",
        "35",
        "36",
        "37",
        "38",
        "39",
        "40",
        "41",
        "42",
        "43",
        "44",
        "45",
        "46",
        "47",
        "48",
        "49",
        "50",
        "51",
        "52",
        "53",
        "54",
        "55",
        "56",
        "57",
        "58",
        "59",
        ), font=Font(family='Helvetica', size=12), width=5,readonlybackground='white')
        self.spinMinAfter.pack(padx=5)
        self.spinAMPMafter = tk.Spinbox(self, values=("","AM","PM"), font=Font(family='Helvetica', size=12), width=5)
        self.spinAMPMafter.pack(padx=5)

        self.labelStartBefore = tk.Label(self, text="FLAG NOTE WITH START TIME BEFORE",font=Font(family='Arial', size=11))
        self.labelStartBefore.pack(fill=tk.X,padx=50)
        self.spinHourBefore = tk.Spinbox(self, values=("","01","02","03","04","05","06","07","08","09","10","11","12"), font=Font(family='Helvetica', size=12), width=5)
        self.spinHourBefore.pack(padx=5)
        self.spinMinBefore = tk.Spinbox(self, values=("","01",
        "02",
        "03",
        "04",
        "05",
        "06",
        "07",
        "08",
        "09",
        "10",
        "11",
        "12",
        "13",
        "14",
        "15",
        "16",
        "17",
        "18",
        "19",
        "20",
        "21",
        "22",
        "23",
        "24",
        "25",
        "26",
        "27",
        "28",
        "29",
        "30",
        "31",
        "32",
        "33",
        "34",
        "35",
        "36",
        "37",
        "38",
        "39",
        "40",
        "41",
        "42",
        "43",
        "44",
        "45",
        "46",
        "47",
        "48",
        "49",
        "50",
        "51",
        "52",
        "53",
        "54",
        "55",
        "56",
        "57",
        "58",
        "59",
        ), font=Font(family='Helvetica', size=12), width=5)
        self.spinMinBefore.pack(padx=5)
        self.spinAMPMbefore = tk.Spinbox(self, values=("","AM","PM"), font=Font(family='Helvetica', size=12), width=5)
        self.spinAMPMbefore.pack(padx=5)


        # Function to flag all individuals under 120 or whatever the user wants to flag

        self.labelUnderUnits = tk.Label(self, text="FLAG INDIVIDUALS WITH TOTAL UNITS LESS THAN:",font=Font(family='Arial', size=11))
        self.labelUnderUnits.pack(fill=tk.X, padx=50)
        self.spinUnderUnits = tk.Entry(self, width=5,font=Font(family='Helvetica', size=12))
        self.spinUnderUnits.pack(padx=5,pady=2)

        
       
        ## Choose file to read
        self.buttonFile = tk.Button(self, text="1. Choose File to be scanned", command=self.file_choose)
        self.buttonFile.pack(fill=tk.X,padx=50,pady=2)
        self.labelFile = tk.Label(self, text="")
        self.labelFile.pack() 

        # Choose direcgtory to save file ot
        self.buttonFileOutput = tk.Button(self, text="2. Confirm output file location", command=self.folder_choose)
        self.buttonFileOutput.pack(fill=tk.X,padx=50,pady=2)
        self.labelFileOutput = tk.Entry(self, text="",background='grey94')
        self.labelFileOutput.pack(fill=tk.X,padx=50,pady=2) 

    

        
        self.button = tk.Button(self, text="3. RUN", command=self.on_button)
        self.button.pack(fill=tk.X,padx=50,pady=2) 
        

        self.labelWorking = tk.Label(self, text="")
        self.labelWorking.pack()

        import configparser
        global config 
        config = configparser.ConfigParser()
        config.read('config.ini')
        self.entryKeywords.insert(0, config.get('DEFAULT', 'entryKeywords'))
        self.spinDurationsGreater.insert(0, config.get('DEFAULT', 'spinDurationsGreater'))
        self.spinDurationsLess.insert(0, config.get('DEFAULT', 'spinDurationsLess'))
        self.spinNoteLength.insert(0, config.get('DEFAULT', 'spinNoteLength'))
        self.spinHourAfter.insert(0, config.get('DEFAULT', 'spinHourAfter'))
        self.spinMinAfter.insert(0, config.get('DEFAULT', 'spinMinAfter'))
        self.spinAMPMafter.insert(0, config.get('DEFAULT', 'spinAMPMafter'))
        self.spinHourBefore.insert(0, config.get('DEFAULT', 'spinHourBefore'))
        self.spinMinBefore.insert(0, config.get('DEFAULT', 'spinMinBefore'))
        self.spinAMPMbefore.insert(0, config.get('DEFAULT', 'spinAMPMbefore'))
        self.spinUnderUnits.insert(0,config.get('DEFAULT','spinUnderUnits'))
        self.labelFileOutput.insert(0,config.get('DEFAULT','labelFileOutput'))
        
        
    #-------------------------  END - MAIN GUI ---------------------------------------------------



    def folder_choose(self):
        global dirname
        dirname = filedialog.askdirectory(parent=self, initialdir="/", title='Please select a directory')
        self.labelFileOutput.delete(0, 'end')
        self.labelFileOutput.insert(0,dirname)
        self.labelFileOutput.pack(fill=tk.X,padx=50)    

    def file_choose(self):
        global file_name
        file_name = filedialog.askopenfilename()
        if file_name.endswith(".xlsx"):
            pass
        else:
            return tk.messagebox.showerror("Warning - File", "Please choose '.xlsx' files only.")
        
        self.labelFile.configure(text=file_name)
        self.labelFile.pack(fill=tk.X,padx=50)
        
    
    def flaggedWords(self, ws, my_list):
        '''Finds keywords in row of data, throws in list'''
        for row in ws.iter_rows(row_offset=1):
            d = row[0] # The note
            e = row[1] # The name of the individual
            f = row[2] # Contact date
            g = row[3] # Program
            h = row[4] # Start time
            i = row[5] # end time
            j = row[6] # duration
            k = row[7] # Note writer
            foundWords = []
            if d.value:
                for w in sorted(my_list):
                    if w.lower() in str(d.value).lower():
                        foundWords.append(w)
                if len(foundWords) > 0:
                    note = ''
                    for l in foundWords:
                        left,sep,right = d.value.lower().partition(l)
                        note = note + left[-70:] + sep.upper() + right[:70] + ';'
                    forCSV = ','.join(foundWords).upper()
                    self.csvWritee(forCSV, e, h, i, f, note, g, j, k)

    def oddDuration(self, ws, greaterthan, lessthan):
        for row in ws.iter_rows(row_offset=1):
            d = row[0] # The note
            e = row[1] # The name of the individual
            f = row[2] # Contact date
            g = row[3] # Program
            h = row[4] # Start time
            i = row[5] # end time
            j = row[6] # duration
            k = row[7] # Note writer
            if d.value:
                if j.value:
                    if j.value <= lessthan or j.value >= greaterthan:
                        note = d.value.split('.')
                        d = '.'.join(note[1:3]).lstrip() + ' [. . .] ' + d.value[-100:]
                        self.csvWritee('Duration', e, h, i, f, d, g, j, k)
                else:
                    self.csvWritee('NO DURATION', e, h, i, f, d.value, g, j, k)
    def shortNote(self, ws, notelength):
        for row in ws.iter_rows(row_offset=1):
            d = row[0] # The note
            e = row[1] # The name of the individual
            f = row[2] # Contact date
            g = row[3] # Program
            h = row[4] # Start time
            i = row[5] # end time
            j = row[6] # duration
            k = row[7] # Note writer
            if d.value:
                if len(d.value) < notelength:
                    self.csvWritee('SHORT NOTE (< ' + str(notelength) + ')', e, h, i, f, d.value, g, j, k)
    def oddTimes(self, ws, startTimeAfter, startTimeBefore):
        after = self.convert24(startTimeAfter)
        afterHour = after[0]
        afterMin = after[1]
        before = self.convert24(startTimeBefore)
        beforeHour = before[0]
        beforeMin = before[1]
        for row in ws.iter_rows(row_offset=1):
            d = row[0] # The note
            e = row[1] # The name of the individual
            f = row[2] # Contact date
            g = row[3] # Program
            h = row[4] # Start time
            i = row[5] # end time
            j = row[6] # duration
            k = row[7] # Note writer
            if h.value:
                note = d.value.split('.')
                d = '.'.join(note[1:3]).lstrip() + ' [. . .] ' + d.value[-100:]
                try:
                    if h.value > time(afterHour, afterMin):
                        self.csvWritee("START TIME AFTER " + startTimeAfter, e, h, i, f, d, g, j, k)
                    elif h.value < time(beforeHour, beforeMin):
                        self.csvWritee("START TIME Before " + startTimeBefore, e, h, i, f, d, g, j, k)
                except (TypeError):
                    self.csvWritee("12AM/Error", e, h, i, f, d, g, j, k)
    def underUnits(self, ws, underUnits):
        units = int(underUnits) * 15
        from collections import defaultdict
        names = defaultdict(int)
        for row in ws.iter_rows(row_offset=1):
            d = row[0] # The note
            e = row[1] # The name of the individual
            f = row[2] # Contact date
            g = row[3] # Program
            h = row[4] # Start time
            i = row[5] # end time
            j = row[6] # duration
            k = row[7] # Note writer

            if j.value is None:
                pass
            else:
                names[e.value] += j.value
        for k, v in names.items():
            if names[k] < units:
                outputWriter.writerow(["UNDER UNITS ("+ str(underUnits) + ')', k, str(int(v)/15)])     
    def convert24(self, str1):
        # Checking if last two elements of time
        # is AM and first two elements are 12
        if str1[-2:] == "AM" and str1[:2] == "12":
            hour = "0"
            minute = str1[3:-3]
            return int(hour), int(minute)   
        # remove the AM    
        elif str1[-2:] == "AM":
            if str1[0] == "0":
                hour = str1[1]
            else:
                hour = str1[:2]
            if str1[3] == "0":
                minute = str1[4]
            else:
                minute = str1[3:-3]
            
            #minute =  str1[3:-3]
            return int(hour), int(minute)
         
        # Checking if last two elements of time
        # is PM and first two elements are 12   
        elif str1[-2:] == "PM" and str1[:2] == "12":
            if str1[0] == "0":
                hour = str1[1]
            else:
                hour = str1[:2]
            if str1[3] == "0":
                minute = str1[4]
            else:
                minute = str1[3:-3]
            return int(hour), int(minute)
             
        else:
             
            # add 12 to hours and remove PM
            hour = int(str1[:2]) + 12
            
            if str1[3] == "0":
                minute = str1[4]
            else:
                minute = str1[3:-3]

            return int(hour), int(minute)
    
    def under7forResidential(self):
        example_dictionary = defaultdict(list)
        for row in ws:
            a = row[0]
            b = row[1]
            if a.value:
                example_dictionary[a.value].append(b.value)

        namesClean = {}

        for names, values in example_dictionary.items():
            namesClean[names] = set(values)

        for names, values in namesClean.items():
            if len(values) < 2:
                print(names + ' has less than two dates')

    
    
    def csvWritee(self, title, e,h,i,f,d,g,j,k):
        
        title = title
        e = e.value
        try:
            h = str(h.value.strftime('%I:%M%p'))
            i = str(i.value.strftime('%I:%M%p'))
        except:
            h,i = "",""
        f = str(f.value.strftime('%m/%d/%Y'))
        d = d
        g = str(g.value)
        j = str(j.value)
        k = str(k.value)
        outputWriter.writerow([title, e, h, i, f, d, g, j, k])

    def on_button(self):
        
        try:
            file_name
        except:
            return tk.messagebox.showerror("Warning - File", "Please choose a file.")

        if self.labelFileOutput.get() == "":
            return  tk.messagebox.showerror("Warning - Directory", "Please choose a save location.") 

        self.saveConfig()

        with open(file_name, "rb") as f:
            in_mem_file = io.BytesIO(f.read())

        trngfile = openpyxl.load_workbook(in_mem_file, read_only=True)
        
        ws = trngfile['Sheet1']
        print(self.labelFileOutput.get())
        outputFileName = str(self.labelFileOutput.get()) + '\AuditCreated-' + str(datetime.datetime.now().strftime('%Y%m%d%H%M%S'))+ '.csv'
        outputFile = open(outputFileName, 'w', newline='')
        global outputWriter
        outputWriter = csv.writer(outputFile)
        outputWriter.writerow(['Flagged Word/Phrase', 'Individual', 'StartTime', 'EndTime', 'Date', 'Excerpt', 'Program', 'Duration','Staff', 'Audit Comments'])

        keywords = self.entryKeywords.get()
        my_list = keywords.split(",")
        print(my_list)
        #if len(my_list) == 1 and my_list[0] == "":
        #    tk.messagebox.showerror("Keywords", "There must be at least one word in the keywords field.")
        #    return
        try:
            greaterthan = int(self.spinDurationsGreater.get())
            lessthan = int(self.spinDurationsLess.get())
            notelength = int(self.spinNoteLength.get())
            unitThreshold = int(self.spinUnderUnits.get())
        except (TypeError, ValueError):
            return tk.messagebox.showerror("Warning - Integer", "Please enter whole numbers only (e.g. 360 or 12)")
        
        startTimeAfter = self.spinHourAfter.get() + ":" + self.spinMinAfter.get() + " " + self.spinAMPMafter.get()
        startTimeBefore = self.spinHourBefore.get() + ":" + self.spinMinBefore.get() + " " + self.spinAMPMbefore.get()
        
        if self.entryKeywords.get().lower() == '':
            pass
        else:
            self.flaggedWords(ws, my_list)
        self.oddDuration(ws, greaterthan, lessthan)
        self.shortNote(ws, notelength)
        self.oddTimes(ws,startTimeAfter,startTimeBefore)
        self.underUnits(ws,unitThreshold)
        self.labelWorking.configure(font=Font(family='Helvetica', size=12),text="FINISHED!")
        outputFile.close()

        def callback(event):
            import os
            import webbrowser
            webbrowser.open_new(r"file://" + os.path.abspath(str(outputFile.name)))
            self.link.configure(text="")

        
        self.link = tk.Label(self, text="Click here for file", fg="blue", cursor="hand2")
        self.link.pack()
        self.link.bind("<Button-1>", callback)
    
    def saveConfig(self):
        config.set('DEFAULT', 'entryKeywords', self.entryKeywords.get())
        config.set('DEFAULT', 'spinDurationsGreater', self.spinDurationsGreater.get())
        config.set('DEFAULT', 'spinDurationsLess', self.spinDurationsLess.get())
        config.set('DEFAULT', 'spinNoteLength', self.spinNoteLength.get())
        config.set('DEFAULT', 'spinHourAfter', self.spinHourAfter.get())
        config.set('DEFAULT', 'spinMinAfter', self.spinMinAfter.get())
        config.set('DEFAULT', 'spinAMPMafter', self.spinAMPMafter.get())
        config.set('DEFAULT', 'spinHourBefore', self.spinHourBefore.get())
        config.set('DEFAULT', 'spinMinBefore', self.spinMinBefore.get())
        config.set('DEFAULT', 'spinAMPMbefore', self.spinAMPMbefore.get())
        config.set('DEFAULT', 'spinUnderUnits', self.spinUnderUnits.get())
        config.set('DEFAULT', 'labelFileOutput', self.labelFileOutput.get())
        config.write(open('config.ini','w'))
        

app = SampleApp()
app.mainloop()